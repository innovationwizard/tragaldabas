from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font

from stages.s8_cell_classification import CellClassifier
from stages.s9_dependency_graph import DependencyGraphBuilder
from stages.s10_logic_extraction import LogicExtractor
from core.models import DependencyGraph, GraphNode
from core.enums import CellRole


@pytest.mark.asyncio
async def test_cell_classification_roles(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["A1"] = "Amount"
    sheet["B1"] = 100
    sheet["C1"] = "=B1*2"
    sheet["D1"] = "=C1+10"
    sheet["A3"] = "Section Header"
    sheet.merge_cells("A3:C3")
    sheet["A3"].font = Font(bold=True)

    file_path = tmp_path / "sample.xlsx"
    workbook.save(file_path)

    classifier = CellClassifier()
    result = await classifier.execute(str(file_path))

    cells = {cell.address: cell for cell in result.sheets[0].cells}

    assert cells["Sheet1!B1"].role.value == "input"
    assert cells["Sheet1!C1"].role.value == "intermediate"
    assert cells["Sheet1!D1"].role.value == "output"
    assert cells["Sheet1!A1"].role.value == "label"
    assert cells["Sheet1!A3"].role.value == "structural"


@pytest.mark.asyncio
async def test_dependency_graph_edges(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["B1"] = 5
    sheet["C1"] = "=B1*2"
    sheet["D1"] = "=C1+1"

    file_path = tmp_path / "graph.xlsx"
    workbook.save(file_path)

    classifier = CellClassifier()
    classification = await classifier.execute(str(file_path))

    builder = DependencyGraphBuilder()
    graph = await builder.execute(classification)

    edges = {(edge.source, edge.target) for edge in graph.edges}
    assert ("Sheet1!B1", "Sheet1!C1") in edges
    assert ("Sheet1!C1", "Sheet1!D1") in edges


@pytest.mark.asyncio
async def test_dependency_graph_clusters_and_depth(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["B1"] = 5
    sheet["C1"] = "=B1*2"
    sheet["D1"] = "=C1+1"

    sheet["F1"] = 3
    sheet["G1"] = "=F1+2"

    file_path = tmp_path / "cluster.xlsx"
    workbook.save(file_path)

    classifier = CellClassifier()
    classification = await classifier.execute(str(file_path))

    builder = DependencyGraphBuilder()
    graph = await builder.execute(classification)

    cluster_ids = {cluster.id for cluster in graph.clusters}
    assert len(cluster_ids) == 2

    node_b1 = graph.nodes["Sheet1!B1"]
    node_c1 = graph.nodes["Sheet1!C1"]
    node_d1 = graph.nodes["Sheet1!D1"]
    node_f1 = graph.nodes["Sheet1!F1"]
    node_g1 = graph.nodes["Sheet1!G1"]

    assert node_b1.depth == 0
    assert node_c1.depth == 1
    assert node_d1.depth == 2
    assert node_f1.depth == 0
    assert node_g1.depth == 1

    assert node_b1.cluster == node_c1.cluster == node_d1.cluster
    assert node_f1.cluster == node_g1.cluster
    assert node_b1.cluster != node_f1.cluster


@pytest.mark.asyncio
async def test_conditional_format_extraction(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = 5
    sheet["A2"] = 15

    rule = FormulaRule(formula=["A1>10"])
    sheet.conditional_formatting.add("A1:A2", rule)

    file_path = tmp_path / "format.xlsx"
    workbook.save(file_path)

    classifier = CellClassifier()
    result = await classifier.execute(str(file_path))

    assert result.conditional_formats


@pytest.mark.asyncio
async def test_logic_extractor_flags_unsupported():
    graph = DependencyGraph(
        nodes={
            "Sheet1!A1": GraphNode(
                address="Sheet1!A1",
                role=CellRole.OUTPUT,
                formula="=INDIRECT(B1)",
            )
        },
        edges=[],
        execution_order=["Sheet1!A1"],
        clusters=[],
        circular_refs=[],
    )
    extractor = LogicExtractor()
    result = await extractor.execute(graph)

    assert result.unsupported_features
