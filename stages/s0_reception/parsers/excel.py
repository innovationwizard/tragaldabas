"""Excel file parser"""

import pandas as pd
from pathlib import Path
from typing import List
import openpyxl

from core.models import (
    ReceptionResult, FileMetadata, SheetPreview
)
from core.enums import FileType
from core.exceptions import FileParseError
from .base import FileParser
from utils.encoding import detect_encoding
from config import settings


class ExcelParser(FileParser):
    """Parser for Excel files (.xlsx, .xls)"""
    
    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls"]
    
    def detect_encoding(self, file_path: str) -> str:
        """Excel files are binary, no encoding needed"""
        return "binary"
    
    def parse(self, file_path: str) -> ReceptionResult:
        """Parse Excel file"""
        path = Path(file_path)
        
        if not path.exists():
            raise FileParseError(f"File not found: {file_path}", file_path)
        
        try:
            file_size = path.stat().st_size
            
            # Determine file type
            if path.suffix.lower() == ".xlsx":
                file_type = FileType.EXCEL_XLSX
            else:
                file_type = FileType.EXCEL_XLS
            
            # Build metadata
            metadata = FileMetadata(
                file_path=str(path.absolute()),
                file_name=path.name,
                file_type=file_type,
                file_size_bytes=file_size,
                encoding=None,
                sheets=[]
            )
            
            # Parse each sheet
            previews = []
            raw_data = {}
            
            if file_type == FileType.EXCEL_XLSX and settings.ETL_INPUTS_ONLY:
                workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
                metadata.sheets = workbook.sheetnames

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    max_row = sheet.max_row or 0
                    max_col = sheet.max_column or 0
                    rows = []
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                        values = []
                        for cell in row:
                            if cell.data_type == "f":
                                values.append("")
                            else:
                                values.append("" if cell.value is None else str(cell.value))
                        rows.append(values)
                    df = pd.DataFrame(rows)

                    col_letters = self._generate_column_letters(len(df.columns))
                    preview_rows = df.head(50).values.tolist()

                    preview = SheetPreview(
                        sheet_name=sheet_name,
                        row_count=len(df),
                        col_count=len(df.columns),
                        preview_rows=preview_rows,
                        column_letters=col_letters
                    )

                    previews.append(preview)
                    raw_data[sheet_name] = df
            else:
                excel_file = pd.ExcelFile(path)
                metadata.sheets = excel_file.sheet_names

                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        dtype=str,
                        keep_default_na=False
                    )
                
                    col_letters = self._generate_column_letters(len(df.columns))
                    preview_rows = df.head(50).values.tolist()

                    preview = SheetPreview(
                        sheet_name=sheet_name,
                        row_count=len(df),
                        col_count=len(df.columns),
                        preview_rows=preview_rows,
                        column_letters=col_letters
                    )

                    previews.append(preview)
                    raw_data[sheet_name] = df
            
            return ReceptionResult(
                metadata=metadata,
                previews=previews,
                raw_data=raw_data
            )
            
        except Exception as e:
            raise FileParseError(
                f"Failed to parse Excel file: {e}",
                file_path
            ) from e
    
    def _generate_column_letters(self, count: int) -> List[str]:
        """Generate Excel-style column letters (A, B, ..., Z, AA, AB, ...)"""
        letters = []
        for i in range(count):
            result = ""
            n = i
            while n >= 0:
                result = chr(n % 26 + ord('A')) + result
                n = n // 26 - 1
            letters.append(result)
        return letters

