"""Stage 8: Cell Classification - interpret Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from core.interfaces import Stage
from core.models import (
    CellClassificationResult,
    ClassifiedCell,
    ConditionalFormat,
    DataValidation,
    NamedRange,
    PivotTableDefinition,
    SheetClassification,
    VBAMacro,
    CellFormatting,
)
from core.enums import CellRole, InputType


class CellClassifier(Stage[str, CellClassificationResult]):
    """Classify workbook cells by role, validation, and references."""

    MAX_RANGE_EXPANSION = 1000

    @property
    def name(self) -> str:
        return "Cell Classification"

    @property
    def stage_number(self) -> int:
        return 8

    def validate_input(self, input_data: str) -> bool:
        if not isinstance(input_data, str):
            return False
        path = Path(input_data)
        return path.exists() and path.is_file()

    async def execute(self, input_data: str) -> CellClassificationResult:
        workbook = load_workbook(input_data, data_only=False, keep_vba=True)
        named_ranges = self._extract_named_ranges(workbook)
        vba_macros = self._extract_vba_macros(input_data, workbook)

        validation_map: Dict[str, DataValidation] = {}
        sheets: List[SheetClassification] = []
        conditional_formats: List[ConditionalFormat] = []
        pivot_tables: List[PivotTableDefinition] = []

        cell_data: Dict[str, Dict[str, object]] = {}
        references_by_cell: Dict[str, Set[str]] = {}
        merged_anchors: Dict[str, Set[Tuple[int, int]]] = {}
        row_stats: Dict[str, Dict[int, Dict[str, int]]] = {}
        row_max_non_empty: Dict[str, int] = {}

        for sheet in workbook.worksheets:
            sheet_title = sheet.title
            validation_map.update(self._extract_validations(workbook, sheet, sheet_title))
            conditional_formats.extend(self._extract_conditional_formats(sheet, sheet_title))
            merged_anchors[sheet_title] = self._extract_merged_anchors(sheet)
            row_stats[sheet_title] = self._build_row_stats(sheet)
            row_max_non_empty[sheet_title] = max(
                (stats["non_empty"] for stats in row_stats[sheet_title].values()),
                default=0,
            )
            pivot_tables.extend(self._extract_pivot_tables(sheet, sheet_title))

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None and cell.data_type != "f":
                        continue

                    address = f"{sheet_title}!{cell.coordinate}"
                    formula = self._normalize_formula(cell.value, cell.data_type)
                    references = []
                    if formula:
                        references = self._extract_references(
                            formula, sheet_title, named_ranges
                        )
                        for ref in references:
                            for expanded in self._expand_reference(ref):
                                references_by_cell.setdefault(expanded, set()).add(address)

                    validation_rule = validation_map.get(address)
                    input_type = self._infer_input_type(cell)
                    if validation_rule:
                        input_type = self._input_type_from_validation(validation_rule) or input_type
                    cell_data[address] = {
                        "value": cell.value,
                        "formula": formula,
                        "references": references,
                        "input_type": input_type,
                        "formatting": self._extract_formatting(cell),
                        "row": cell.row,
                        "col": cell.column,
                        "sheet": sheet_title,
                    }

        # Ensure referenced cells exist even if empty in workbook
        for target_address in list(references_by_cell.keys()):
            if target_address not in cell_data:
                cell_data[target_address] = {
                    "value": None,
                    "formula": None,
                    "references": [],
                    "input_type": InputType.TEXT,
                    "formatting": None,
                }

        sheet_cells: Dict[str, List[ClassifiedCell]] = {}
        role_by_coord: Dict[str, Dict[Tuple[int, int], CellRole]] = {}
        structural_rows: Dict[str, List[Tuple[int, str]]] = {}

        for address, payload in cell_data.items():
            sheet_name, _ = address.split("!", 1)
            referenced_by = sorted(references_by_cell.get(address, set()))
            formula = payload.get("formula")

            if formula:
                role = CellRole.INTERMEDIATE if referenced_by else CellRole.OUTPUT
            elif referenced_by:
                role = CellRole.INPUT
            else:
                role = CellRole.STATIC

            role_by_coord.setdefault(sheet_name, {})[(payload["row"], payload["col"])] = role

            classified = ClassifiedCell(
                address=address,
                role=role,
                input_type=payload.get("input_type"),
                formula=formula,
                value=payload.get("value"),
                validation=validation_map.get(address),
                formatting=payload.get("formatting"),
                referenced_by=referenced_by,
                references=payload.get("references") or [],
            )

            sheet_cells.setdefault(sheet_name, []).append(classified)

        for sheet_name, cells in sheet_cells.items():
            sheet_structural: List[Tuple[int, str]] = []
            for cell in cells:
                if cell.role != CellRole.STATIC:
                    continue
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                coord = self._parse_coordinate(cell.address)
                if not coord:
                    continue

                if self._is_structural_cell(
                    sheet_name,
                    coord,
                    cell,
                    merged_anchors,
                    row_stats,
                    row_max_non_empty.get(sheet_name, 0),
                ):
                    cell.role = CellRole.STRUCTURAL
                elif self._is_label_cell(sheet_name, coord, role_by_coord):
                    cell.role = CellRole.LABEL
                if cell.role in {CellRole.STRUCTURAL, CellRole.LABEL}:
                    cell.label = str(cell.value).strip()
                if cell.role == CellRole.STRUCTURAL:
                    sheet_structural.append((coord[0], cell.label or "Section"))

            structural_rows[sheet_name] = sorted(sheet_structural)
            sheets.append(
                SheetClassification(
                    name=sheet_name,
                    cells=sorted(cells, key=lambda c: c.address),
                    input_groups=self._build_groups(sheet_name, cells, structural_rows),
                    output_groups=self._build_groups(
                        sheet_name, cells, structural_rows, role_filter=CellRole.OUTPUT
                    ),
                    sections=self._build_sections(sheet_name, cells, structural_rows),
                )
            )

        return CellClassificationResult(
            sheets=sheets,
            named_ranges=named_ranges,
            vba_macros=vba_macros,
            data_validations=list(validation_map.values()),
            conditional_formats=conditional_formats,
            pivot_tables=pivot_tables,
        )

    def _normalize_formula(self, value: object, data_type: Optional[str]) -> Optional[str]:
        if data_type != "f":
            return None
        if value is None:
            return None
        formula = str(value)
        if not formula.startswith("="):
            formula = f"={formula}"
        return formula

    def _extract_named_ranges(self, workbook) -> List[NamedRange]:
        named_ranges: List[NamedRange] = []
        for defined_name in workbook.defined_names.values():
            if not defined_name.name:
                continue
            for sheet_name, ref in defined_name.destinations:
                named_ranges.append(NamedRange(name=defined_name.name, ref=f"{sheet_name}!{ref}"))
        return named_ranges

    def _extract_validations(
        self, workbook, sheet, sheet_title: str
    ) -> Dict[str, DataValidation]:
        validations: Dict[str, DataValidation] = {}
        if not sheet.data_validations:
            return validations

        for validation in sheet.data_validations.dataValidation:
            rule_type = getattr(validation, "type", "") or ""
            formula1 = getattr(validation, "formula1", None)
            options = self._parse_validation_options(workbook, sheet_title, rule_type, formula1)

            for cell in validation.cells:
                address = f"{sheet_title}!{cell}"
                validations[address] = DataValidation(
                    address=address,
                    validation_type=rule_type,
                    operator=getattr(validation, "operator", None),
                    formula1=formula1,
                    formula2=getattr(validation, "formula2", None),
                    allow_blank=bool(getattr(validation, "allowBlank", False)),
                    options=options,
                    error_message=getattr(validation, "error", None),
                    prompt_title=getattr(validation, "promptTitle", None),
                    prompt_message=getattr(validation, "prompt", None),
                )

        return validations

    def _extract_references(
        self,
        formula: str,
        sheet_name: str,
        named_ranges: List[NamedRange],
    ) -> List[str]:
        references: List[str] = []
        named_map = {nr.name: nr.ref for nr in named_ranges}
        try:
            tokenizer = Tokenizer(formula)
            for token in tokenizer.items:
                if token.subtype in {"RANGE", "CELL", "NAMED_RANGE"}:
                    value = token.value.replace("$", "")
                    if token.subtype == "NAMED_RANGE" and value in named_map:
                        references.append(self._normalize_reference(named_map[value], sheet_name))
                    else:
                        references.append(self._normalize_reference(value, sheet_name))
        except Exception:
            pass

        return [ref for ref in references if ref]

    def _normalize_reference(self, ref: str, default_sheet: str) -> str:
        ref = ref.replace("$", "")
        if "!" in ref:
            sheet_name, address = ref.split("!", 1)
            sheet_name = sheet_name.strip("'")
            return f"{sheet_name}!{address}"
        return f"{default_sheet}!{ref}"

    def _expand_reference(self, ref: str) -> Iterable[str]:
        if "!" not in ref:
            return []
        sheet_name, address = ref.split("!", 1)
        if ":" not in address:
            return [f"{sheet_name}!{address}"]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(address)
        except ValueError:
            return []

        total = (max_row - min_row + 1) * (max_col - min_col + 1)
        if total > self.MAX_RANGE_EXPANSION:
            return [f"{sheet_name}!{address}"]

        expanded = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                expanded.append(f"{sheet_name}!{self._col_letter(col)}{row}")
        return expanded

    def _col_letter(self, col_idx: int) -> str:
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _infer_input_type(self, cell) -> InputType:
        if cell.is_date:
            return InputType.DATE
        value = cell.value
        number_format = (cell.number_format or "").lower()
        if isinstance(value, bool):
            return InputType.BOOLEAN
        if isinstance(value, (int, float)):
            if "%" in number_format:
                return InputType.PERCENTAGE
            if "$" in number_format or "€" in number_format or "£" in number_format:
                return InputType.CURRENCY
            return InputType.NUMBER
        return InputType.TEXT

    def _extract_formatting(self, cell) -> CellFormatting:
        font_color = None
        fill_color = None
        if cell.font and cell.font.color:
            font_color = getattr(cell.font.color, "rgb", None)
        if cell.fill and getattr(cell.fill, "fgColor", None):
            fill_color = getattr(cell.fill.fgColor, "rgb", None)

        return CellFormatting(
            number_format=cell.number_format,
            font_bold=bool(getattr(cell.font, "bold", False)),
            font_italic=bool(getattr(cell.font, "italic", False)),
            font_color=font_color,
            fill_color=fill_color,
        )

    def _extract_conditional_formats(
        self, sheet, sheet_title: str
    ) -> List[ConditionalFormat]:
        formats: List[ConditionalFormat] = []
        cf_rules = getattr(sheet.conditional_formatting, "_cf_rules", {}) or {}
        for range_str, rules in cf_rules.items():
            for rule in rules:
                formulas = getattr(rule, "formula", None)
                if isinstance(formulas, list):
                    formula_text = ",".join(formulas)
                else:
                    formula_text = str(formulas) if formulas else ""
                rule_type = getattr(rule, "type", "rule")
                rule_desc = f"{rule_type}:{formula_text}".strip(":")
                format_type, format_color, severity = self._format_from_rule(rule)
                formats.append(
                    ConditionalFormat(
                        range=f"{sheet_title}!{range_str}",
                        rule=rule_desc,
                        format_type=format_type,
                        format_color=format_color,
                        severity=severity,
                    )
                )
        return formats

    def _format_from_rule(self, rule) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        dxf = getattr(rule, "dxf", None)
        if not dxf:
            return None, None, None
        color = None
        if getattr(dxf, "fill", None) and getattr(dxf.fill, "fgColor", None):
            color = getattr(dxf.fill.fgColor, "rgb", None)
        if not color and getattr(dxf, "font", None) and getattr(dxf.font, "color", None):
            color = getattr(dxf.font.color, "rgb", None)
        severity = None
        if color:
            upper = color.upper()
            if upper.startswith("FF") and upper.endswith("0000"):
                severity = "error"
            elif "FF0000" in upper:
                severity = "error"
            elif "FFA500" in upper or "FFCC00" in upper:
                severity = "warning"
            elif "00FF00" in upper:
                severity = "info"
        return getattr(rule, "type", None), color, severity

    def _extract_pivot_tables(
        self, sheet, sheet_title: str
    ) -> List[PivotTableDefinition]:
        pivots: List[PivotTableDefinition] = []
        raw_pivots = getattr(sheet, "_pivots", None) or []
        for pivot in raw_pivots:
            name = getattr(pivot, "name", None) or getattr(pivot, "cache", None) or "pivot"
            source = getattr(getattr(pivot, "cache", None), "ref", None) or ""
            rows = []
            cols = []
            values = []
            filters = []
            try:
                row_fields = getattr(pivot, "rowFields", None)
                if row_fields:
                    rows = [str(field) for field in getattr(row_fields, "field", [])]
                col_fields = getattr(pivot, "colFields", None)
                if col_fields:
                    cols = [str(field) for field in getattr(col_fields, "field", [])]
                data_fields = getattr(pivot, "dataFields", None)
                if data_fields:
                    values = [str(field) for field in getattr(data_fields, "dataField", [])]
                page_fields = getattr(pivot, "pageFields", None)
                if page_fields:
                    filters = [str(field) for field in getattr(page_fields, "pageField", [])]
            except Exception:
                rows = []
            pivots.append(
                PivotTableDefinition(
                    name=str(name),
                    source_range=f"{sheet_title}!{source}" if source else sheet_title,
                    rows=rows,
                    columns=cols,
                    values=values,
                    filters=filters,
                )
            )
        return pivots

    def _extract_vba_macros(self, file_path: str, workbook) -> List[VBAMacro]:
        macros: List[VBAMacro] = []
        archive = getattr(workbook, "vba_archive", None)
        if not archive:
            return macros
        try:
            from oletools.olevba import VBA_Parser
        except Exception:
            macros.append(
                VBAMacro(
                    name="vbaProject",
                    code="VBA project detected, but oletools is not available.",
                )
            )
            return macros

        parser = VBA_Parser(file_path)
        try:
            if not parser.detect_vba_macros():
                return macros
            for (_, _, filename, content) in parser.extract_macros():
                if not content:
                    continue
                macros.append(
                    VBAMacro(
                        name=str(filename),
                        code=content,
                    )
                )
        finally:
            parser.close()
        return macros

    def _parse_validation_options(
        self,
        workbook,
        sheet_title: str,
        validation_type: str,
        formula1: Optional[str],
    ) -> List[str]:
        if validation_type != "list" or not formula1:
            return []
        formula = formula1.strip()
        if formula.startswith('"') and formula.endswith('"'):
            formula = formula[1:-1]
        if "," in formula:
            return [item.strip() for item in formula.split(",") if item.strip()]
        if formula.startswith("="):
            formula = formula[1:]
        if "!" in formula:
            sheet_name, ref = formula.split("!", 1)
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = sheet_title
            ref = formula
        try:
            target_sheet = workbook[sheet_name]
        except Exception:
            return []
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except ValueError:
            return []
        values: List[str] = []
        for row in target_sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value).strip()
                if value:
                    values.append(value)
        return values
        return []

    def _input_type_from_validation(self, validation: DataValidation) -> Optional[InputType]:
        if validation.validation_type == "list":
            return InputType.ENUM
        if validation.validation_type in {"whole", "decimal"}:
            return InputType.NUMBER
        if validation.validation_type == "date":
            return InputType.DATE
        if validation.validation_type == "textLength":
            return InputType.TEXT
        return None

    def _build_sections(
        self,
        sheet_name: str,
        cells: List[ClassifiedCell],
        structural_rows: Dict[str, List[Tuple[int, str]]],
    ) -> List[SheetSection]:
        sections: List[SheetSection] = []
        if not structural_rows.get(sheet_name):
            return sections
        row_to_section: Dict[int, str] = {}
        for row_idx, label in structural_rows.get(sheet_name, []):
            row_to_section[row_idx] = label

        boundaries = self._section_boundaries(structural_rows.get(sheet_name, []))
        for cell in cells:
            coord = self._parse_coordinate(cell.address)
            if not coord:
                continue
            section = self._find_section(sheet_name, coord, structural_rows)
            if not section:
                continue
            if not self._coord_in_section(coord, section, boundaries):
                continue
            section_entry = next((s for s in sections if s.name == section), None)
            if not section_entry:
                section_entry = SheetSection(name=section, cells=[])
                sections.append(section_entry)
            section_entry.cells.append(cell.address)
        return sections

    def _build_groups(
        self,
        sheet_name: str,
        cells: List[ClassifiedCell],
        structural_rows: Dict[str, List[Tuple[int, str]]],
        role_filter: CellRole = CellRole.INPUT,
    ) -> List[InputGroup | OutputGroup]:
        groups: Dict[str, List[str]] = {}
        boundaries = self._section_boundaries(structural_rows.get(sheet_name, []))
        for cell in cells:
            if cell.role != role_filter:
                continue
            coord = self._parse_coordinate(cell.address)
            if not coord:
                continue
            section = self._find_section(sheet_name, coord, structural_rows) or "General"
            if section != "General" and not self._coord_in_section(coord, section, boundaries):
                continue
            groups.setdefault(section, []).append(cell.address)
        def format_name(section: str) -> str:
            suffix = "Outputs" if role_filter == CellRole.OUTPUT else "Inputs"
            if section == "General":
                return f"{sheet_name} - General {suffix}"
            return f"{sheet_name} - {section} {suffix}"

        if role_filter == CellRole.OUTPUT:
            return [
                OutputGroup(name=format_name(key), cells=values)
                for key, values in groups.items()
            ]
        return [
            InputGroup(name=format_name(key), cells=values)
            for key, values in groups.items()
        ]

    def _find_section(
        self,
        sheet_name: str,
        coord: Tuple[int, int],
        structural_rows: Dict[str, List[Tuple[int, str]]],
    ) -> Optional[str]:
        row, _ = coord
        for struct_row, label in reversed(structural_rows.get(sheet_name, [])):
            if struct_row <= row and label:
                return label
        return None

    def _section_boundaries(
        self, rows: List[Tuple[int, str]]
    ) -> Dict[str, Tuple[int, Optional[int]]]:
        boundaries: Dict[str, Tuple[int, Optional[int]]] = {}
        if not rows:
            return boundaries
        sorted_rows = sorted(rows, key=lambda item: item[0])
        for idx, (row_idx, label) in enumerate(sorted_rows):
            start = row_idx + 1
            end = sorted_rows[idx + 1][0] - 1 if idx + 1 < len(sorted_rows) else None
            boundaries[label] = (start, end)
        return boundaries

    def _coord_in_section(
        self,
        coord: Tuple[int, int],
        section: str,
        boundaries: Dict[str, Tuple[int, Optional[int]]],
    ) -> bool:
        row, _ = coord
        if section not in boundaries:
            return False
        start, end = boundaries[section]
        if end is None:
            return row >= start
        return start <= row <= end

    def _extract_merged_anchors(self, sheet) -> Set[Tuple[int, int]]:
        anchors: Set[Tuple[int, int]] = set()
        for merged in sheet.merged_cells.ranges:
            anchors.add((merged.min_row, merged.min_col))
        return anchors

    def _build_row_stats(self, sheet) -> Dict[int, Dict[str, int]]:
        counts: Dict[int, Dict[str, int]] = {}
        for row in sheet.iter_rows():
            non_empty = 0
            text_count = 0
            numeric_count = 0
            row_idx = row[0].row if row else 0
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    non_empty += 1
                    if isinstance(cell.value, str):
                        text_count += 1
                    elif isinstance(cell.value, (int, float)):
                        numeric_count += 1
            counts[row_idx] = {
                "non_empty": non_empty,
                "text": text_count,
                "numeric": numeric_count,
            }
        return counts

    def _parse_coordinate(self, address: str) -> Optional[Tuple[int, int]]:
        if "!" not in address:
            return None
        _, coord = address.split("!", 1)
        try:
            return coordinate_to_tuple(coord)
        except ValueError:
            return None

    def _is_structural_cell(
        self,
        sheet_name: str,
        coord: Tuple[int, int],
        cell: ClassifiedCell,
        merged_anchors: Dict[str, Set[Tuple[int, int]]],
        row_stats: Dict[str, Dict[int, Dict[str, int]]],
        max_non_empty: int,
    ) -> bool:
        if coord in merged_anchors.get(sheet_name, set()):
            return True
        if cell.formatting and cell.formatting.font_bold:
            return True
        stats = row_stats.get(sheet_name, {}).get(coord[0], {})
        non_empty = stats.get("non_empty", 0)
        text_count = stats.get("text", 0)
        numeric_count = stats.get("numeric", 0)
        if non_empty <= 1:
            return True
        if text_count >= 2 and numeric_count == 0:
            return True
        if max_non_empty and non_empty >= max(3, int(max_non_empty * 0.7)):
            if text_count >= max(2, int(non_empty * 0.8)):
                return True
        if isinstance(cell.value, str):
            value = cell.value.strip()
            if value.endswith(":"):
                return True
            if value.isupper() and len(value) >= 4:
                return True
            if value.lower().startswith(("total", "summary", "subtotal")):
                return True
        return False

    def _is_label_cell(
        self,
        sheet_name: str,
        coord: Tuple[int, int],
        role_by_coord: Dict[str, Dict[Tuple[int, int], CellRole]],
    ) -> bool:
        neighbors = [
            (coord[0], coord[1] - 1),
            (coord[0], coord[1] + 1),
            (coord[0] - 1, coord[1]),
            (coord[0] + 1, coord[1]),
        ]
        for neighbor in neighbors:
            role = role_by_coord.get(sheet_name, {}).get(neighbor)
            if role in {CellRole.INPUT, CellRole.OUTPUT, CellRole.INTERMEDIATE}:
                return True
        return False
